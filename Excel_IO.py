import xlrd
import openpyxl
import openpyxl.styles as bea
book = xlrd.open_workbook("income.xlsx")
sheet = book.sheet_by_name("2018")
print(f"{book.nsheets}")
print(f"{book.sheet_names()}")
print(sheet.ncols)
print(sheet.nrows)
print(sheet.cell_value(rowx=0,colx=0))
print(sheet.row_values(rowx=0))
money = sheet.col_values(colx=1,start_rowx=1)
print(sum(money))

book2 = openpyxl.Workbook()
sh = book2.active
sh.title = 'aaa'
sh2 = book2.create_sheet('asd',0)
sh3 = book2.create_sheet('ads',1)
sh = book2['aaa']
sh['A1']='a'
sh['A1'].font=bea.Font(color=bea.colors.BLUE,size=15,bold=True)
print(sh['A1'].value)
sh.cell(2,2).value='?'
book2.save('1.xlsx')